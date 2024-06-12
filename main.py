import glob
import os.path
import pickle
import time

import fuzzysearch
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side

import utils

from structure import *
from layout_parser import contour_parser, rule_parser
from reader import paddleocr_reader, vietocr_reader
from content_parser import kor_parser


def read_image(img, page_idx=0):
    page = Page(img, page_idx=page_idx)
    res_line = paddleocr_reader.detect_line(img)
    for l in res_line[0]:
        img_line, _ = utils.crop_4_points(img, l)
        line_obj = Line(l[0][0], l[0][1], l[1][0], l[1][1], l[2][0], l[2][1], l[3][0], l[3][1], name='line')
        line_obj.img = img_line
        page.lines.append(line_obj)

    texts, probs = vietocr_reader.read_batch([l.img for l in page.lines])
    for l, t, p in zip(page.lines, texts, probs):
        l.content = t
        l.prob = p
    utils.remove_lower_prob_lines(page, prob_threshold=0.3)
    if len(page.lines) == 0:
        return None
    rule_parser.parse_layout(page)

    # find keyword to break document
    content = str(page.text).strip()
    return content


def page_process(img, page_idx=0, debug=False) -> Page:
    page = Page(img, page_idx=page_idx)
    res_line = paddleocr_reader.detect_line(img)
    for l in res_line[0]:
        img_line, _ = utils.crop_4_points(img, l)
        line_obj = Line(l[0][0], l[0][1], l[1][0], l[1][1], l[2][0], l[2][1], l[3][0], l[3][1], name='line')
        line_obj.img = img_line
        page.lines.append(line_obj)

    texts, probs = vietocr_reader.read_batch([l.img for l in page.lines])
    for l, t, p in zip(page.lines, texts, probs):
        l.content = t
        l.prob = p
    utils.remove_lower_prob_lines(page, prob_threshold=0.3)
    if len(page.lines) == 0:
        return None
    rule_parser.parse_layout(page)
    if debug:
        utils.plot_block(page, level=1, line=True)

    # find keyword to break document
    content = str(page.text).strip()
    query_strings = ['cộng hoà xã hội chủ nghĩa việt nam', 'Độc lập - Tự do - Hạnh phúc', 'Mẫu số ', "DANH SÁCH "]
    for q in query_strings:
        matches = fuzzysearch.find_near_matches(q.lower(), content.lower()[:1000],
                                                max_l_dist=int(len(q) * 0.3))
        if matches:
            matches = list(sorted(matches, key=lambda x: x.dist))
            head = content[matches[0].end:750]
            title = kor_parser.extract_title(head)
            page.title = title
            page.sub_doc_anchor = True
            break

    return page


def split_document(document: Document):
    document.sub_docs = []
    start_page_idx = 0
    pre_title = None

    for i, page in enumerate(document.pages):
        if page.sub_doc_anchor:
            if pre_title is None:
                # đầu văn bản, chưa có văn bản trước đó
                pre_title = page.title
            else:
                if page.title == pre_title:
                    # 1 văn bản có 2 trang có title
                    pass
                else:
                    # sang văn bản tiếp theo
                    sub_doc = SubDocument()
                    sub_doc.start_page_idx = start_page_idx
                    sub_doc.end_page_idx = i
                    sub_doc.pages = document.pages[start_page_idx: sub_doc.end_page_idx]
                    sub_doc.title = pre_title
                    sub_doc.idx = len(document.sub_docs)
                    document.sub_docs.append(sub_doc)

                    pre_title = page.title
                    start_page_idx = i

    sub_doc = SubDocument()
    sub_doc.start_page_idx = start_page_idx
    sub_doc.end_page_idx = len(document.pages)
    sub_doc.pages = document.pages[start_page_idx: sub_doc.end_page_idx]
    sub_doc.title = pre_title
    sub_doc.idx = len(document.sub_docs)
    document.sub_docs.append(sub_doc)

    return document


def main_process(pdf_path: str, output_path="data/output/", debug=False):
    print(f"Process file: {pdf_path}")
    pdf_name = os.path.basename(pdf_path)
    pdf_id, ext = os.path.splitext(pdf_name)

    output_path = os.path.join(output_path, pdf_id)
    if not os.path.exists(output_path):
        os.makedirs(output_path, exist_ok=True)

    output_dump = os.path.join(output_path, f'{pdf_id}.pkl')
    if os.path.exists(output_dump) and not debug:
        print(f"File {pdf_path} has processed.")
        return pickle.load(open(output_dump, 'rb'))

    start_time = time.time()
    document = Document(pdf_path)

    images = utils.get_image_from_pdf2(pdf_path=pdf_path)
    print(f"{pdf_id}: {len(images)} pages")
    idx = 0
    for i, img in enumerate(images):
        page = page_process(img, idx)
        if page:
            document.pages.append(page)
            idx += 1

    split_document(document)

    pickle.dump(document, open(output_dump, 'wb'))

    i = 0
    for sub_doc in document.sub_docs:
        output_pdf_split = os.path.join(output_path, f'{pdf_id}_{i}.pdf')
        utils.insert_image_to_pdf([x.img for x in sub_doc.pages], output_pdf_split)
        i += 1

    output_text = os.path.join(output_path, f'{pdf_id}.txt')
    with open(output_text, 'w') as f:
        for idx, page in enumerate(document.pages):
            f.write(str(page.text))
            f.write(f'\n\n====================={idx}\n\n')

    output_excel = os.path.join(output_path, f'{pdf_id}.xlsx')
    export_excel(document, os.path.join(output_path, f'{pdf_id}_'), output_excel)

    print(f"Time process: {time.time() - start_time:.4f}s")

    return document


def export_excel(document: Document, prefix_out_pdf, output_file):
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove_sheet(default_sheet)
    # =========== create sheet tailieudinhkem =================
    sub_doc_sheet = workbook.create_sheet("tailieudinhkem", 1)
    headers = ["Mã hồ sơ \n1", "Tên tài liệu\n2", "Số hiệu\n3", "Thời gian\n4", "Số thửa\n5", "Số tờ bản đồ\n6",
               "Diện tích\n7", "Loại đất\n8", "Tác giả\n9", "Trích yếu\n10\n", "Tờ số\n11", "Mô tả/Ghi chú\n12",
               "File đính kèm (PDF)\n13"]
    # create header
    header_style = NamedStyle(name="header")
    header_style.font = Font(bold=True)
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_style.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                                 bottom=Side(style="thin"), )
    for col_idx, header in enumerate(headers, start=1):
        cell = sub_doc_sheet.cell(row=1, column=col_idx, value=header)
        cell.style = header_style

    # fill data
    for row_idx, sub_doc in enumerate(document.sub_docs, start=2):
        cell = sub_doc_sheet.cell(row=row_idx, column=1, value=document.file_id)
        cell = sub_doc_sheet.cell(row=row_idx, column=2, value=sub_doc.title)
        cell = sub_doc_sheet.cell(row=row_idx, column=12, value="Bản chính")
        cell = sub_doc_sheet.cell(row=row_idx, column=13, value=f"{prefix_out_pdf}{sub_doc.idx}.pdf")

    # ============= create sheet hosotailieu ============
    doc_sheet = workbook.create_sheet("hosotailieu", 0)
    headers = ["Mã hồ sơ\n1", "Mã đơn vị hành chính\n2", "Mục lục tư liệu\n 3", "Loại tư liệu \n4", "Tên tư liệu \n5",
               "Địa chỉ người sử dụng\n6", "Phông số \n7", "Địa chỉ đơn vị hình thành Phông\n8",
               "Thời gian bắt đầu \n9",
               "Thời gian kết thúc\n 10", "Vị trí", "Thời hạn bảo quản\n15", "Đơn vị bảo quản\n16", "Số lượng\n17",
               "Đơn vị tính\n18", "Ngôn ngữ TL\n19", "Chế độ sử dụng TL\n20", "Độ mật\n21", "Nơi lưu trữ\n22",
               "Địa chỉ lưu trữ\n23", "Ghi chú\n24"]
    # create header
    col_idx = 1
    for header in headers:
        cell = doc_sheet.cell(row=1, column=col_idx, value=header)
        cell.style = header_style
        if col_idx == 2:
            cell = doc_sheet.cell(row=2, column=2, value="Mã tỉnh")
            cell.style = header_style
            cell = doc_sheet.cell(row=2, column=3, value="Mã huyện")
            cell.style = header_style
            cell = doc_sheet.cell(row=2, column=4, value="Mã xã")
            cell.style = header_style
            col_idx += 2
        elif col_idx == 13:
            cell = doc_sheet.cell(row=2, column=13, value="Kho/Phòng\n11")
            cell.style = header_style
            cell = doc_sheet.cell(row=2, column=14, value="Giá/Tủ\n12")
            cell.style = header_style
            cell = doc_sheet.cell(row=2, column=15, value="Ngăn\n13")
            cell.style = header_style
            cell = doc_sheet.cell(row=2, column=16, value="Hộp\n14")
            cell.style = header_style
            col_idx += 3
        else:
            doc_sheet.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        col_idx += 1
    doc_sheet.merge_cells('B1:D1')
    doc_sheet.merge_cells('M1:P1')

    # fill data
    cell = doc_sheet.cell(row=3, column=1, value=document.file_id)
    cell = doc_sheet.cell(row=3, column=2, value=10)
    cell = doc_sheet.cell(row=3, column=3, value="BacGiang")
    cell = doc_sheet.cell(row=3, column=5, value="DAT_DAI")
    cell = doc_sheet.cell(row=3, column=17, value="Vĩnh viễn")
    cell = doc_sheet.cell(row=3, column=18, value="Văn phòng Đăng ký đất đai tỉnh Lào Cai")
    cell = doc_sheet.cell(row=3, column=19, value=1)
    cell = doc_sheet.cell(row=3, column=20, value="Hồ sơ")
    cell = doc_sheet.cell(row=3, column=21, value="Tiếng việt")
    cell = doc_sheet.cell(row=3, column=23, value="Thường")
    cell = doc_sheet.cell(row=3, column=24, value="Văn phòng Đăng ký đất đai tỉnh Lào Cai")
    cell = doc_sheet.cell(row=3, column=25, value="Khối 7, Phường Nam Cường, Thành phố Lào Cai, Tỉnh Lào Cai")

    workbook.save(filename=output_file)
    print(f"Save excel to file {output_file}")


def process_get_title_from_pkl():
    for path_pkl in glob.glob("data/dump/*.pkl"):
        print(path_pkl)
        # path_pkl = "data/dump/XG_01035-11a.pkl"
        document = pickle.load(open(path_pkl, 'rb'))
        for sub in document.sub_docs:
            if sub.title:
                print(f"Has sub title: {sub.title}")
                continue
            content = str(sub.text).strip()
            head = content[:800]
            print(head)
            title = kor_parser.extract_title(head)
            print("=========")
            print(title)
            sub.title = title

            pickle.dump(document, open(path_pkl, 'wb'))


if __name__ == '__main__':
    pdf_paths = glob.glob('/Users/tienthien/Downloads/zalo/DUNG_PhanMem/ThucHien/*.pdf')
    files = ['XG_01116-2.pdf', 'XG_01025-10.pdf', 'XG_01032-15.pdf']
    pdf_paths = [f'/Users/tienthien/Downloads/zalo/DUNG_PhanMem/ThucHien/{f}' for f in files]
    # pdf_paths = ['/Volumes/thien/data_tc/data_raw/SongMai_TPBG/k co trong vilis/SM_00001.pdf']
    for path in pdf_paths:
        main_process(path, debug=True)
    # file_pdf = 'XG_01117-1'
    # document = pickle.load(open(f"data/output/{file_pdf}/{file_pdf}.pkl", 'rb'))
    # print(document)
    # split_document(document)
    #
    # export_excel(document, f'data/output/{file_pdf}/{file_pdf}.xlsx')
